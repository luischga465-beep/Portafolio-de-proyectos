
#FASE 1
import openpyxl
from copy import copy

# Definir los nombres y rutas de acceso de los archivos mensuales
archivos_mensuales = {
    "M1": r"C:\Users\kavo6\Downloads\1 Ingresos de enero 2023- 2024.xlsx",
    "M2": r"C:\Users\kavo6\Downloads\1 Ingresos de febrero 2023- 2024.xlsx",
    "M3": r"C:\Users\kavo6\Downloads\1. Ingresos de marzo 2023- 2024.xlsx"
}

Periodo = "Ene-Mar"
NumTri = "Primer Trimestre"
NT = "1er"
años = "2023-2024"
año1 = "2023"
año2 = "2024"
Mes1 = "Enero"
Mes2 = "Febrero"
Mes3 = "Marzo"

# Convertir el formato de los archivos originales
def replace_formulas_with_values(file_paths):
    for file_path in file_paths:
        try:
            # Abre el archivo Excel existente
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Recorre todas las hojas del archivo
            for sheet in workbook.worksheets:
                # Recorre todas las celdas de la hoja
                for row in sheet.iter_rows():
                    for cell in row:
                        # Verifica si la celda tiene una fórmula
                        if cell.value is not None:
                            # Obtiene el valor resultante de la fórmula
                            value = cell.value
                            # Reemplaza la fórmula con el valor resultante
                            cell.value = value

            # Guarda el archivo modificado
            workbook.save(file_path)
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Error: El archivo {file_path} no es un archivo Excel válido o tiene un formato no soportado.")

# Extrae las rutas de los archivos del diccionario y pásalas a la función
replace_formulas_with_values(archivos_mensuales.values())

# Definir el nombre y la ruta de acceso del archivo trimestral
archivo_trimestral = r"C:\Users\kavo6\Downloads\Informe Trimestral.xlsx"

# Crear un nuevo libro de Excel para el archivo trimestral
libro_trimestral = openpyxl.Workbook()

# Iterar sobre cada archivo mensual
for mes, ruta in archivos_mensuales.items():
    # Abrir el archivo mensual
    libro_mensual = openpyxl.load_workbook(ruta)
    # Obtener la primera hoja del archivo mensual
    primera_hoja = libro_mensual.active
    # Crear una nueva hoja en el archivo trimestral con el nombre del mes
    hoja_trimestral = libro_trimestral.create_sheet(title=primera_hoja.title)

    # Copiar datos y formato celda por celda
    for fila in primera_hoja.iter_rows(values_only=True):
        hoja_trimestral.append(fila)
    
    for row in primera_hoja.iter_rows(min_row=1, max_row=primera_hoja.max_row, min_col=1, max_col=primera_hoja.max_column):
        for cell in row:
            new_cell = hoja_trimestral[cell.coordinate]
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.alignment = copy(cell.alignment)
            new_cell.protection = copy(cell.protection)

    # Copiar tamaño de las columnas y filas
    for col in primera_hoja.columns:
        hoja_trimestral.column_dimensions[col[0].column_letter].width = primera_hoja.column_dimensions[col[0].column_letter].width
    for row in primera_hoja.rows:
        hoja_trimestral.row_dimensions[row[0].row].height = primera_hoja.row_dimensions[row[0].row].height

# Eliminar la hoja inicial que se crea automáticamente al inicio
libro_trimestral.remove(libro_trimestral["Sheet"])

# Añadir la nueva hoja de ingresos trimestrales
hoja_ingresos_trim = libro_trimestral.create_sheet(title=f"Ingresos {NT} Trim {años}")

# Copiar toda la información de la primera hoja al nuevo ingreso trimestral
for fila in libro_trimestral[libro_trimestral.sheetnames[0]].iter_rows(values_only=True):
    hoja_ingresos_trim.append(fila)

for row in libro_trimestral[libro_trimestral.sheetnames[0]].iter_rows(min_row=1, max_row=libro_trimestral[libro_trimestral.sheetnames[0]].max_row, min_col=1, max_col=libro_trimestral[libro_trimestral.sheetnames[0]].max_column):
    for cell in row:
        new_cell = hoja_ingresos_trim[cell.coordinate]
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.alignment = copy(cell.alignment)
        new_cell.protection = copy(cell.protection)

# Guardar los cambios en el archivo trimestral
libro_trimestral.save(archivo_trimestral)


# FASE 2
# Obtener la hoja de ingresos trimestrales
hoja_ingresos_trim = libro_trimestral[f"Ingresos {NT} Trim {años}"]

# Obtener la primera hoja del archivo trimestral
primera_hoja_trimestral = libro_trimestral[libro_trimestral.sheetnames[0]]

# Copiar anchos de columnas
for col in primera_hoja_trimestral.columns:
    hoja_ingresos_trim.column_dimensions[col[0].column_letter].width = primera_hoja_trimestral.column_dimensions[col[0].column_letter].width

# Copiar alturas de filas
for row in primera_hoja_trimestral.rows:
    hoja_ingresos_trim.row_dimensions[row[0].row].height = primera_hoja_trimestral.row_dimensions[row[0].row].height

# Iterar sobre todas las celdas de la hoja
for fila in hoja_ingresos_trim.iter_rows():
    for celda in fila:
        # Verificar si el valor de la celda es un número
        if isinstance(celda.value, (int, float)):
            # Eliminar el valor de la celda
            celda.value = None

# Obtener las hojas correspondientes a los meses de manera dinámica
hojas_meses = libro_trimestral.sheetnames[:3]  # Asumiendo que las primeras tres hojas son las correspondientes a los meses
hojas_meses_objs = [libro_trimestral[sheet] for sheet in hojas_meses]

# Iterar sobre las filas y columnas de las hojas de los meses
for fila_trim, *filas_meses in zip(hoja_ingresos_trim.iter_rows(), *map(lambda hoja: hoja.iter_rows(), hojas_meses_objs)):
    for celda_trim, *celdas_meses in zip(fila_trim, *filas_meses):
        # Sumar las cantidades correspondientes de cada celda para cada mes
        if all(isinstance(celda.value, (int, float)) for celda in celdas_meses):
            celda_trim.value = sum(celda.value for celda in celdas_meses)


# Modificar los valores en la primera columna "Costo por Inmovilizador" (columna H)
for fila in hoja_ingresos_trim.iter_rows(min_row=2, min_col=8, max_col=8):
    for celda in fila:
        if celda.value == 882:
            celda.value = 294

# Modificar los valores en la segunda columna "Costo por Inmovilizador" (columna S)
for fila in hoja_ingresos_trim.iter_rows(min_row=2, min_col=19, max_col=19):
    for celda in fila:
        if celda.value == 936:
            celda.value = 312

# Reemplazar el texto en la celda D38
hoja_ingresos_trim["D38"] = f"Total {NumTri} {año1}"

# Reemplazar el texto en la celda O38
hoja_ingresos_trim["O38"] = f"Total {NumTri} {año2}"

# Reemplazar el texto en la celda B2
hoja_ingresos_trim["B2"] = f"{Mes1} - {Mes3} {año1}"

# Reemplazar el texto en la celda B2
hoja_ingresos_trim["M2"] = f"{Mes1} - {Mes3} {año2}"

# Guardar los cambios en el archivo trimestral
libro_trimestral.save(archivo_trimestral)


# FASE 3
# Crear una nueva hoja llamada Ajustes
hoja_ajustes_2024 = libro_trimestral.create_sheet(title=f"Ajustes {año2}")

# Obtener la fila 3 de la hoja de ingresos trimestrales desde la columna N
fila_encabezados = [celda.value for celda in hoja_ingresos_trim[3] if celda.column_letter >= 'N']

# Pegar los encabezados en la hoja de ajustes 2024 a partir de la celda C4
for i, encabezado in enumerate(fila_encabezados, start=3):
    hoja_ajustes_2024.cell(row=4, column=i, value=encabezado)

# Agregar la palabra "Mes" en la celda B4
hoja_ajustes_2024['B4'] = 'Mes'

# Copiar los encabezados faltantes (columnas AA y AB)
hoja_ajustes_2024['Q4'] = hoja_ingresos_trim['AB3'].value

from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment

# Ajustar el alto de la fila 4 a 60.75 y el ancho de las columnas a 20 a partir de la columna C
hoja_ajustes_2024.row_dimensions[4].height = 60.75
for col in range(hoja_ajustes_2024.min_column, hoja_ajustes_2024.max_column + 1):
    if col >= 3:
        letra_columna = get_column_letter(col)
        hoja_ajustes_2024.column_dimensions[letra_columna].width = 20

# Alinear y centrar el texto en cada celda de la fila 4
for col in range(2, hoja_ajustes_2024.max_column + 1):
    letra_columna = get_column_letter(col)
    celda = hoja_ajustes_2024[f"{letra_columna}4"]
    
    # Crear un objeto de alineación
    alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Aplicar la alineación a la celda
    celda.alignment = alineacion

# Copiar los valores de las variables en la hoja "Ajustes 2024"
meses = [Mes1, Mes2, Mes3]
fila_inicio = 5  # La fila donde comenzará a copiarse los valores

for idx, mes in enumerate(meses):
    fila_inicio_mes = fila_inicio + 7 * idx  # Añadir una fila en blanco antes de cada valor de mes
    celda_inicio = hoja_ajustes_2024.cell(row=fila_inicio_mes, column=2)  # Comenzar en la fila correspondiente
    for i in range(6):  # Copiar el valor del mes seis veces hacia abajo
        celda_inicio.offset(row=i, column=0).value = mes
    
    # Agregar el texto "Total 'Mes'" en la celda en blanco después del valor de mes
    celda_total = celda_inicio.offset(row=6, column=0)
    celda_total.value = f"Total {mes}"

# Guardar los cambios en el archivo trimestral
libro_trimestral.save(archivo_trimestral)

#AAAAAAAAAAAAAAAAAAAAAAAAA
# Obtener la primera hoja y la hoja "Ajustes 2024"
hoja_origen = libro_trimestral[libro_trimestral.sheetnames[0]]  # La primera hoja

#hoja_ajustes_2024 = libro_trimestral["Ajustes 2024"]

# Filas y columnas a copiar
filas_a_copiar = [14, 16, 19, 24, 26, 32, 36]
columna_inicio = 14  # Columna N

# Fila de inicio en la hoja "Ajustes 2024"
fila_inicio_ajustes = 5

# Copiar y pegar los valores
for i, fila_idx in enumerate(filas_a_copiar):
    for col_idx in range(columna_inicio, hoja_origen.max_column + 1):
        celda_origen = hoja_origen.cell(row=fila_idx, column=col_idx)
        celda_destino = hoja_ajustes_2024.cell(row=fila_inicio_ajustes + i, column=col_idx - columna_inicio + 3)
        celda_destino.value = celda_origen.value

# Sumar los valores de las primeras dos filas y poner el resultado en la fila siguiente
fila_suma = fila_inicio_ajustes  # Fila donde se pondrá el resultado

for col_idx in range(columna_inicio, hoja_origen.max_column + 1):
    celda_1 = hoja_ajustes_2024.cell(row=fila_inicio_ajustes, column=col_idx - columna_inicio + 3)
    celda_2 = hoja_ajustes_2024.cell(row=fila_inicio_ajustes + 1, column=col_idx - columna_inicio + 3)

    # Verificar si ambas celdas son números
    if isinstance(celda_1.value, (int, float)) and isinstance(celda_2.value, (int, float)):
        celda_1.value = celda_1.value + celda_2.value
    else:
        celda_1.value = "Polanco - Anzures"

# Mover hacia arriba los elementos para llenar el espacio vacío
for fila in range(7, hoja_ajustes_2024.max_row + 1):  # Comenzar desde la fila 7
    for col in range(3, hoja_ajustes_2024.max_column + 1):  # Columnas a partir de C
        hoja_ajustes_2024.cell(row=fila - 1, column=col).value = hoja_ajustes_2024.cell(row=fila, column=col).value

# Borrar la última fila para eliminar los datos duplicados
for col in range(3, hoja_ajustes_2024.max_column + 1):  # Columnas a partir de C
    hoja_ajustes_2024.cell(row=hoja_ajustes_2024.max_row, column=col).value = None

# Sumar los valores de las filas 5 a 10 en la fila 11
for col_idx in range(3, hoja_ajustes_2024.max_column + 1):  # Iterar sobre las columnas desde la columna C
    total = 0  # Inicializar el total para cada columna
    for fila_idx in range(5, 11):  # Iterar sobre las filas 5 a 10
        celda = hoja_ajustes_2024.cell(row=fila_idx, column=col_idx)
        # Verificar si la celda contiene un número
        if isinstance(celda.value, (int, float)):
            total += celda.value  # Sumar el valor de la celda al total
        else:
            hoja_ajustes_2024.cell(row=11, column=col_idx).value = None  # Dejar la celda en blanco si no es un número
            break  # Salir del bucle interno si encontramos una celda que no es un número
    else:
        hoja_ajustes_2024.cell(row=11, column=col_idx).value = total  # Colocar el total en la fila 11 si todas las celdas son números

#BBBBBBBBBBBBBB
hoja_origen = libro_trimestral[libro_trimestral.sheetnames[1]]  # La segunda hoja

# Filas y columnas a copiar
filas_a_copiar = [14, 16, 19, 24, 26, 32, 36]
columna_inicio = 14  # Columna N

# Fila de inicio en la hoja "Ajustes 2024"
fila_inicio_ajustes = 12

# Copiar y pegar los valores
for i, fila_idx in enumerate(filas_a_copiar):
    for col_idx in range(columna_inicio, hoja_origen.max_column + 1):
        celda_origen = hoja_origen.cell(row=fila_idx, column=col_idx)
        celda_destino = hoja_ajustes_2024.cell(row=fila_inicio_ajustes + i, column=col_idx - columna_inicio + 3)
        celda_destino.value = celda_origen.value

# Sumar los valores de las primeras dos filas y poner el resultado en la fila siguiente
fila_suma = fila_inicio_ajustes  # Fila donde se pondrá el resultado

for col_idx in range(columna_inicio, hoja_origen.max_column + 1):
    celda_1 = hoja_ajustes_2024.cell(row=fila_inicio_ajustes, column=col_idx - columna_inicio + 3)
    celda_2 = hoja_ajustes_2024.cell(row=fila_inicio_ajustes + 1, column=col_idx - columna_inicio + 3)

    # Verificar si ambas celdas son números
    if isinstance(celda_1.value, (int, float)) and isinstance(celda_2.value, (int, float)):
        celda_1.value = celda_1.value + celda_2.value
    else:
        celda_1.value = "Polanco - Anzures"

# Mover hacia arriba los elementos para llenar el espacio vacío
for fila in range(14, hoja_ajustes_2024.max_row + 1):  # Comenzar desde la fila 14
    for col in range(3, hoja_ajustes_2024.max_column + 1):  # Columnas a partir de C
        hoja_ajustes_2024.cell(row=fila - 1, column=col).value = hoja_ajustes_2024.cell(row=fila, column=col).value

# Borrar la última fila para eliminar los datos duplicados
for col in range(3, hoja_ajustes_2024.max_column + 1):  # Columnas a partir de C
    hoja_ajustes_2024.cell(row=hoja_ajustes_2024.max_row, column=col).value = None


# Sumar los valores de las filas 12 a 17 en la fila 18
for col_idx in range(3, hoja_ajustes_2024.max_column + 1):  # Iterar sobre las columnas desde la columna C
    total = 0  # Inicializar el total para cada columna
    for fila_idx in range(12, 18):  # Iterar sobre las filas 12 a 17
        celda = hoja_ajustes_2024.cell(row=fila_idx, column=col_idx)
        # Verificar si la celda contiene un número
        if isinstance(celda.value, (int, float)):
            total += celda.value  # Sumar el valor de la celda al total
        else:
            hoja_ajustes_2024.cell(row=18, column=col_idx).value = None  # Dejar la celda en blanco si no es un número
            break  # Salir del bucle interno si encontramos una celda que no es un número
    else:
        hoja_ajustes_2024.cell(row=18, column=col_idx).value = total  # Colocar el total en la fila 18 si todas las celdas son números

#CCCCCCCCCCCCCCCCCCCCC
hoja_origen = libro_trimestral[libro_trimestral.sheetnames[2]]  # La segunda hoja

# Filas y columnas a copiar
filas_a_copiar = [14, 16, 19, 24, 26, 32, 36]
columna_inicio = 14  # Columna N

# Fila de inicio en la hoja "Ajustes 2024"
fila_inicio_ajustes = 19

# Copiar y pegar los valores
for i, fila_idx in enumerate(filas_a_copiar):
    for col_idx in range(columna_inicio, hoja_origen.max_column + 1):
        celda_origen = hoja_origen.cell(row=fila_idx, column=col_idx)
        celda_destino = hoja_ajustes_2024.cell(row=fila_inicio_ajustes + i, column=col_idx - columna_inicio + 3)
        celda_destino.value = celda_origen.value

# Sumar los valores de las primeras dos filas y poner el resultado en la fila siguiente
fila_suma = fila_inicio_ajustes  # Fila donde se pondrá el resultado

for col_idx in range(columna_inicio, hoja_origen.max_column + 1):
    celda_1 = hoja_ajustes_2024.cell(row=fila_inicio_ajustes, column=col_idx - columna_inicio + 3)
    celda_2 = hoja_ajustes_2024.cell(row=fila_inicio_ajustes + 1, column=col_idx - columna_inicio + 3)

    # Verificar si ambas celdas son números
    if isinstance(celda_1.value, (int, float)) and isinstance(celda_2.value, (int, float)):
        celda_1.value = celda_1.value + celda_2.value
    else:
        celda_1.value = "Polanco - Anzures"

# Mover hacia arriba los elementos para llenar el espacio vacío
for fila in range(21, hoja_ajustes_2024.max_row + 1):  # Comenzar desde la fila 21
    for col in range(3, hoja_ajustes_2024.max_column + 1):  # Columnas a partir de C
        hoja_ajustes_2024.cell(row=fila - 1, column=col).value = hoja_ajustes_2024.cell(row=fila, column=col).value

# Borrar la última fila para eliminar los datos duplicados
for col in range(3, hoja_ajustes_2024.max_column + 1):  # Columnas a partir de C
    hoja_ajustes_2024.cell(row=hoja_ajustes_2024.max_row, column=col).value = None

# Sumar los valores de las filas 19 a 24 en la fila 25
for col_idx in range(3, hoja_ajustes_2024.max_column + 1):  # Iterar sobre las columnas desde la columna C
    total = 0  # Inicializar el total para cada columna
    for fila_idx in range(19, 25):  # Iterar sobre las filas 19 a 24
        celda = hoja_ajustes_2024.cell(row=fila_idx, column=col_idx)
        # Verificar si la celda contiene un número
        if isinstance(celda.value, (int, float)):
            total += celda.value  # Sumar el valor de la celda al total
        else:
            hoja_ajustes_2024.cell(row=25, column=col_idx).value = None  # Dejar la celda en blanco si no es un número
            break  # Salir del bucle interno si encontramos una celda que no es un número
    else:
        hoja_ajustes_2024.cell(row=25, column=col_idx).value = total  # Colocar el total en la fila 25 si todas las celdas son números

# Eliminar la columna D
hoja_ajustes_2024.delete_cols(4)

filas_a_sumar = [11, 18, 25]
fila_destino = 27

# Columna de inicio (columna D corresponde a la columna 4)
columna_inicio = 4

# Sumar los valores de las filas 11, 18 y 25 y poner el resultado en la fila 27
for col_idx in range(columna_inicio, hoja_ajustes_2024.max_column + 1):
    total = 0
    for fila_idx in filas_a_sumar:
        celda = hoja_ajustes_2024.cell(row=fila_idx, column=col_idx)
        if isinstance(celda.value, (int, float)):  # Verificar si la celda contiene un número
            total += celda.value
    hoja_ajustes_2024.cell(row=fila_destino, column=col_idx).value = total

# Columna G (columna 7)
columna_g = 7

# Iterar sobre todas las filas de la columna G
for fila in range(1, hoja_ajustes_2024.max_row + 1):
    celda = hoja_ajustes_2024.cell(row=fila, column=columna_g)
    if isinstance(celda.value, (int, float)):  # Verificar si la celda contiene un número
        celda.value = 312

# Limpiar el contenido de las columnas K y O
for row in range(1, hoja_ajustes_2024.max_row + 1):
    hoja_ajustes_2024.cell(row=row, column=11).value = None  # Columna K es la columna 11
    hoja_ajustes_2024.cell(row=row, column=15).value = None  # Columna O es la columna 15

hoja_ajustes_2024['D3'] = 'A'
hoja_ajustes_2024['E3'] = 'B'
hoja_ajustes_2024['F3'] = 'C'
hoja_ajustes_2024['G3'] = 'D'
hoja_ajustes_2024['H3'] = 'E'
hoja_ajustes_2024['I3'] = 'F'
hoja_ajustes_2024['J3'] = 'G'
hoja_ajustes_2024['L3'] = 'H'
hoja_ajustes_2024['M3'] = 'I'
hoja_ajustes_2024['N3'] = 'J'
hoja_ajustes_2024['P3'] = 'G+H+I+J'

# Guardar el libro
libro_trimestral.save(archivo_trimestral)


#FASE 4
# Añadir una nueva hoja
hoja_transacciones_trim = libro_trimestral.create_sheet(title=f"Transacciones {Periodo} {años}")

# Inicializar la fila inicial para pegar los datos
fila_inicial = 2

# Iterar sobre cada archivo mensual
for mes, ruta in archivos_mensuales.items():
    # Abrir el archivo mensual
    libro_mensual = openpyxl.load_workbook(ruta)
    # Obtener la segunda hoja del archivo mensual
    hoja_mensual = libro_mensual.worksheets[1]

    # Copiar datos y formato celda por celda
    for fila in hoja_mensual.iter_rows(values_only=True):
        hoja_transacciones_trim.append(fila)

# Definir las filas a eliminar
filas_a_eliminar = [12, 13, 14, 24, 25, 26]

# Eliminar las filas y desplazar el contenido hacia arriba
for fila in sorted(filas_a_eliminar, reverse=True):
    hoja_transacciones_trim.delete_rows(fila)

# Definir el rango de filas para copiar información basada en texto
fila_inicio = 30
fila_origen_inicio = 21
fila_origen_fin = 29  # 9 filas arriba de 30

# Iterar sobre el rango de columnas
for col in range(1, hoja_transacciones_trim.max_column + 1):
    col_letter = get_column_letter(col)
    
    for fila in range(fila_origen_inicio, fila_origen_fin + 1):
        celda_origen = hoja_transacciones_trim.cell(row=fila, column=col)
        celda_destino = hoja_transacciones_trim.cell(row=fila + 9, column=col)
        
        # Copiar solo si la celda contiene texto
        if isinstance(celda_origen.value, str):
            celda_destino.value = celda_origen.value

# Asignar los valores a las celdas especificadas
hoja_transacciones_trim['L2'] = f"Variación {años}"
hoja_transacciones_trim['L3'] = Mes1
hoja_transacciones_trim['L12'] = Mes2
hoja_transacciones_trim['L21'] = Mes3
hoja_transacciones_trim['L30'] = f"{NT} Trimestre"
hoja_transacciones_trim['B38'] = f"Total Transacciones {NT} Trimestre"
hoja_transacciones_trim['G38'] = f"Total Transacciones {NT} Trimestre"

fila_inicio = 31
fila_fin = 37
columnas = ['B', 'G']

# Iterar sobre el rango de filas y columnas
for fila in range(fila_inicio, fila_fin + 1):
    for col in columnas:
        hoja_transacciones_trim[f'{col}{fila}'] = f"{NT} Trimestre"

# Definir el rango de filas para las columnas E y J
fila_inicio = 31
columnas = ['E', 'J']

# Función para calcular la suma específica
def calcular_suma_desde(fila_inicial, columna):
    suma = 0
    fila_actual = fila_inicial
    while True:
        valor_actual = hoja_transacciones_trim[f'{columna}{fila_actual}'].value
        if valor_actual is None:  # Si encontramos una celda vacía, terminamos
            break
        if isinstance(valor_actual, (int, float)):  # Solo sumamos si es un número
            suma += valor_actual
        fila_actual += 9  # Saltar 9 filas
    return suma

# Iterar sobre las filas a partir de fila_inicio
fila_referencia = 4  # Fila inicial de la suma
for fila in range(fila_inicio, hoja_transacciones_trim.max_row + 1):
    for columna in columnas:
        suma = calcular_suma_desde(fila_referencia, columna)  # Calcular la suma a partir de la fila de referencia
        hoja_transacciones_trim[f'{columna}{fila}'].value = suma
    fila_referencia += 1  # Mover una fila hacia abajo

# Definir el rango de filas a iterar
fila_inicio = 2  # Comienza en la fila 2 si la fila 1 tiene encabezados
fila_fin = hoja_transacciones_trim.max_row  # Hasta la última fila con datos

# Iterar sobre las filas para calcular la variación porcentual
for fila in range(fila_inicio, fila_fin + 1):
    celda_E = hoja_transacciones_trim[f'E{fila}'].value
    celda_J = hoja_transacciones_trim[f'J{fila}'].value

    if isinstance(celda_E, (int, float)) and isinstance(celda_J, (int, float)):
        if celda_E != 0:  # Evitar la división por cero
            variacion_porcentual = ((celda_J - celda_E) / celda_E)
            hoja_transacciones_trim[f'L{fila}'].value = variacion_porcentual
        else:
            hoja_transacciones_trim[f'L{fila}'].value = None  # Evitar división por cero

#Cambiar el ancho de columnas
# Recorrer cada columna
for col in range(1, hoja_transacciones_trim.max_column + 1):
    col_letter = get_column_letter(col)
    adjust_width = False
    
    # Recorrer cada fila de la columna
    for row in range(1, hoja_transacciones_trim.max_row + 1):
        cell = hoja_transacciones_trim.cell(row=row, column=col)
        
        # Verificar si la celda tiene contenido
        if cell.value is not None and cell.value != '':
            adjust_width = True
            break
    
    # Ajustar el ancho de la columna si se encontró contenido en alguna celda
    if adjust_width:
        hoja_transacciones_trim.column_dimensions[col_letter].width = 20

         
# Guardar los cambios en el archivo trimestral
libro_trimestral.save(archivo_trimestral)
